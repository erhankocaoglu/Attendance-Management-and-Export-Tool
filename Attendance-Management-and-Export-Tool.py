import tkinter as tk
from tkinter import ttk, filedialog
import openpyxl


class FileProcess:
    def __init__(self):
        self.filename = "" # prevents an error

    def excel_file_selection(self, pick): # opens file explorer
        self.filename = filedialog.askopenfilename(title="Select Excel File",
                                              filetypes=(("Excel Files", "*.xlsx"), ("All Files", "*.*")))

        if self.filename:
            pick(self.filename)



def save_excel_file(self, input_file, output_file):
    try:
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active

        output_wb = openpyxl.Workbook()
        output_ws = output_wb.active

        row_number = 1

        for row in ws.iter_rows(values_only=True):
            column_number = 1

            for cell in row:
                output_ws.cell(row=row_number, column=column_number, value=cell)
                column_number += 1

                # Creates a loop in order to prevent exceeden more columns than we need
                if column_number % 4 == 0:
                    row_number += 1
                    column_number = 1  # Makes program go back to first column

        output_wb.save(output_file)

        print("Excel file has written sucessfully")
    except Exception as e:
        print(f"ERROR!! : {e}")


class Manager:
    def __init__(self, master, fileprocess):
        self.master = master
        self.fileprocess = fileprocess
        self.current_excel_file = None # Prevents an error(making it equal to None)
        master.title("Project 1")


        # ALL OF THE WIDGETS

        label = tk.Label(master, text="  AttendenceKeeper v1.0", font=("Helvetica", 20, "bold"))
        label.grid(row=0, column=0, columnspan=3, pady=(10, 0))

        label = tk.Label(master, text="Please select a file type ", font=("Helvetica", 10, "bold"))
        label.grid(row=3, column=0, pady=(0), padx=90)

        self.filetype = ttk.Combobox(master, values=[".txt", ".csv", ".xls"], width=10, height=20,
                                     font=("Helvetica", 10))
        self.filetype.grid(row=3, column=0, padx=0, pady=0, sticky="e")

        self.combobox = ttk.Combobox(master,
                                     values=["AP 01", "AP 02", "AP 03", "AP 04", "AP 05", "AP 06", "AP 07", "AP 08",
                                             "AP 09", "AP 10", "AP 11", "AP 12", "AP 13", "AP 14", "AP 15", "AP 16",
                                             "AP 17", "AP 18", "AP 19", "AP 20"], width=12, font=("Helvetica", 8),
                                     height=15)
        self.combobox.grid(row=4, column=0, pady=0, sticky="e")
        self.combobox.bind("<<ComboboxSelected>>", self.section_selected)

        label = tk.Label(master, text="Please select a section:", font=("Helvetica", 10, "bold"))
        label.grid(row=4, column=0, pady=0, padx=0, sticky="")

        self.frame = tk.Frame(master)
        self.frame.grid(row=3, column=1, pady=5, sticky="nsew")

        self.listbox_left = tk.Listbox(master, width=5, height=5, font=("Helvetica", 8))
        self.listbox_left.grid(row=2, column=0, pady=0, padx=0, sticky="nsew")

        self.listbox_right = tk.Listbox(master, width=5, height=5, font=("Helvetica", 8))
        self.listbox_right.grid(row=2, column=2, pady=0, padx=0, sticky="nsew")

        button_width = 8
        button_height = 1

        self.button1 = tk.Button(master, text="Import", font=("Helvetica", 8), command=self.call_excel_file,
                                 width=button_width,
                                 height=button_height)
        self.button1.grid(row=1, column=1, pady=0, sticky="ew")

        label = tk.Label(master, text="     Select student list excel file  :", font=("Helvetica", 10, "bold"))
        label.grid(row=1, column=0, columnspan=1, pady=(0), sticky="e")

        label = tk.Label(master, text="Attended students  :", font=("Helvetica", 10, "bold"))
        label.grid(row=1, column=2, columnspan=1, pady=(0), sticky="w")

        self.button2 = tk.Button(master, text="Add", font=("Helvetica", 8), command=self.button2_attribute,
                                 width=button_width, height=button_height)
        self.button2.grid(row=2, column=1, pady=5, sticky="n")

        self.button3 = tk.Button(master, text="Delete", font=("Helvetica", 8), command=self.button3_attribute,
                                 width=button_width, height=button_height)
        self.button3.grid(row=2, column=1, pady=50, sticky="n")

        self.button4 = tk.Button(master, text="Export", font=("Helvetica", 8), command=self.button4_attribute,
                                 width=button_width, height=button_height)
        self.button4.grid(row=3, column=3, columnspan=1, padx=(0), pady=(0, 0), sticky="w")

        entry_width = 30

        label = tk.Label(master, text="File's name ", font=("Helvetica", 10, "bold"))
        label.grid(row=3, column=2, padx=120, pady=(0, 0), sticky="w")

        self.entry1 = tk.Entry(master, font=("Helvetica", 8))
        self.entry1.grid(row=3, column=2, pady=0, sticky="e")

        master.grid_rowconfigure(0, weight=1)
        master.grid_rowconfigure(1, weight=1)
        master.grid_rowconfigure(1, weight=1)
        master.grid_rowconfigure(2, weight=1)
        master.grid_rowconfigure(3, weight=1)
        master.grid_columnconfigure(0, weight=1)
        master.grid_columnconfigure(1, weight=1)
        master.grid_columnconfigure(2, weight=1)

    def call_excel_file(self): # connects excel_file_selection to load_excel_file
        self.fileprocess.excel_file_selection(self.load_excel_file)
    # load the excel file that is imported and sends it current_excel_file
    def load_excel_file(self, filename):
        self.current_excel_file = filename
        self.master.title(f"Project 1 - {filename}")
    # Assigns student names to left listbox
    def process_and_insert_names(self, filename):
        try:
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active

            selected_section = self.combobox.get()
            for row in sheet.iter_rows(values_only=True):
                if str(row[3]) == selected_section:
                    self.listbox_left.insert(tk.END, row[1])

            workbook.close()
        except Exception as e:
            print(f"Error!! : {e}")
    # Filter left listbox according to the section that is selected currently
    def section_selected(self, event):
        selected_section = self.combobox.get()
        self.listbox_left.delete(0, tk.END)

        if self.current_excel_file:
            self.process_and_insert_names(self.current_excel_file)
        else:
            print("Please select an excel file first.")
    # ADD FUNCTİON
    def button2_attribute(self):
        selected_item = self.listbox_left.get(tk.ACTIVE)
        self.listbox_left.delete(tk.ACTIVE)
        self.listbox_right.insert(tk.END, selected_item)
    # DELETE FUNCTION
    def button3_attribute(self):
        selected_item = self.listbox_right.get(tk.ACTIVE)
        self.listbox_right.delete(tk.ACTIVE)
        self.listbox_left.insert(tk.END, selected_item)
    # EXPORT FUNCTION
    def button4_attribute(self):
        selected_items = self.listbox_right.get(0, tk.END)
        if not selected_items:
            print("Please select at least a student first.")
            return

        file_extension = self.filetype.get()

        if not file_extension:
            print("Please select a file extension first.")
            return
        # CSV FILE
        if file_extension == ".csv":
            raise BaseException("File type is not supported.")
        elif file_extension == ".txt":  # TXT FILE

            selected_section = self.combobox.get()  # Filters according to section
            file_name = f"{selected_section} {self.entry1.get()}{file_extension}"  # Creates a file name
            file_path = filedialog.asksaveasfilename(defaultextension=file_extension, initialfile=file_name)

            if not file_path:
                print("File could not be saved.")
                return

            try:
                wb = openpyxl.load_workbook(self.current_excel_file)
                ws = wb.active

                with open(file_path, 'w', encoding='utf-8') as txt_file:
                    txt_file.write("ID\tName\t           Dept.\n")  # For txt file it adds ID Name Dept top of the file

                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[1] in selected_items:
                            txt_file.write(f"{row[0]}\t{row[1]} {row[2]}\t\n") # ADDS first(ID) second(Name) third(Dept.)
                    print("File saved succesfully.")
            except Exception as e:
                print(f"Error while saving file : {e}")


        else:  # Else is for EXCEL FILE

            selected_section = self.combobox.get()  # Filters according to section that is selected

            file_name = f"{selected_section} {self.entry1.get()}{file_extension}"  # File name

            file_path = filedialog.asksaveasfilename(defaultextension=file_extension, initialfile=file_name)

            if not file_path:
                print("File could not be saved.")

                return

            try:

                wb = openpyxl.load_workbook(self.current_excel_file)

                ws = wb.active

                # CREATES EMPTY LIST

                filtered_rows = []

                # Copies the first row(I did that in order to add the titles(NAME ID DEPT SECTION))

                filtered_rows.append(next(ws.iter_rows(values_only=True)))

                # Takes selected names and send them to filtered_rows

                for row in ws.iter_rows(min_row=2, values_only=True):

                    if row[1] in selected_items:
                        filtered_row = row[:3]  # Until fourth row(fourth is not included)

                        filtered_rows.append(filtered_row)

                # Opens a completely new excel file

                output_wb = openpyxl.Workbook()

                output_ws = output_wb.active

                # Copies the first three columns that we have filtered

                for row_index, row_data in enumerate(filtered_rows, start=1):

                    for col_index, cell_data in enumerate(row_data, start=1):
                        output_ws.cell(row=row_index, column=col_index, value=cell_data)

                # Saves the file

                output_wb.save(file_path)

                print("File saved succesfully.")

            except Exception as e:

                print(f"Error while file saving: {e}")


class Main:
    def __init__(self):
        root = tk.Tk()
        self.fileprocess = FileProcess()
        self.manager = Manager(root, self.fileprocess)
        root.mainloop()


if __name__ == "__main__":
    Main()

